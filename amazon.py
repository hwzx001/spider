from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import pandas as pd
from multiprocessing.dummy import Pool
import re
import time
import random
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment

# Get product links related to keywords
class SPIDER():
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('lang=en_US')
        self.driver=webdriver.Chrome(options=chrome_options)
        self.driver.get(r'https://www.amazon.com')
    def __del__(self):
        self.driver.close()

    def scroll(self):
        # Pull the scroll bar all the way to load the web page
        for i in range(2, 1000):
            js = "var q=document.documentElement.scrollTop={}".format(i * 100)  # javascript
            self.driver.execute_script(js)

    def send_keys(self,keys):
        self.driver.find_element_by_id('twotabsearchtextbox').send_keys(keys)
        self.driver.find_element_by_class_name('nav-input').click()
        try:
            self.scroll()
            pages_location=self.driver.find_element_by_class_name('a-pagination')
            pages=pages_location.find_elements_by_tag_name('a')
            res=[]
            for i in pages:
                res.append(i.get_attribute('href'))
            if  len(res)<=3:
                res=res[0:len(res)-1]
                return res
            else:
                last_page = int(self.driver.find_element_by_xpath(
                    '//*[@id="search"]/div[1]/div[2]/div/span[8]/div/span/div/div/ul/li[6]').text)
                raw_url=res[1]
                return parse_raw_url(raw_url=raw_url,last_page=last_page)
        except:
            return []

    def get_shop_url(self,pageurl):
        res=[]
        self.driver.get(pageurl)
        self.scroll()
        #//*[@id="search"]/div[1]/div[2]/div/span[4]
        product_session = self.driver.find_element_by_xpath('//*[@id="search"]/div[1]/div[2]/div/span[4]/div[1]').find_elements_by_class_name('a-link-normal')
        pat=re.compile('/dp/[0-9A-Z]+/')
        cmp_no=[]
        pat2 = re.compile(r'ref=sr_(.+)\?')
        for item in product_session:
            temp=item.get_attribute('href')
            '''if re.search(pat,temp)!=None:
                res.append(temp)'''
            if  'offer' not in temp and 'customer'not in temp and temp not in res and re.search(pat,temp)!=None:
                s=re.search(pat2,temp)
                if s!=None :
                    no=s.group(1)
                    if no not in cmp_no:
                        cmp_no.append(s.group(1))
                        res.append(temp)
        return res

    def get_allshop_url(self,keys):
        try:
            pagelist=self.send_keys(keys)
        except:
            return []
        else:
            res=[]
            for url in pagelist:
                try:
                    self.driver.get(url)
                    self.scroll()
                    temp=self.get_shop_url(url)
                except:
                    pass
                else:
                    for i in temp:
                        res.append(i)
            return res
# To get one shop information
class SHOPSPIDER():
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('lang=en_US')
        self.driver = webdriver.Chrome(options=chrome_options)

    def init_one_shop(self,url):
        self.driver.get(url)
        self.scroll()

    def scroll(self):
        # Pull the scroll bar all the way to load the web page
        for i in range(2, 1000):
            js = "var q=document.documentElement.scrollTop={}".format(i * 100)  # javascript
            self.driver.execute_script(js)

    def get_title(self):
        try:
            title=self.driver.find_element_by_id('productTitle')
        except:
            return ''
        else:
            return str(title.text).strip()

    def get_score(self):
        try:
            score=str(self.driver.find_element_by_id('acrPopover').get_attribute('title'))[0:3]
        except:
            return ''
        else :
            return score

    def get_price_and_delivery(self):
        try:
            raw_price=str(self.driver.find_element_by_id('price').text).strip()[7:]
        except:
            return ''
        else:
            return raw_price
        '''
        p1=re.compile(r'\$([0-9.]+) ')
        p2=re.compile(r'\+(\w)+ ')
        price=re.search(p1,raw_price).group(1)
        shiping=re.search(p2,raw_price).group(1)
        return price,shiping'''

    def get_list_item(self):
        res=[]
        try:
            list_item=self.driver.find_element_by_xpath('//*[@id="feature-bullets"]/ul').find_elements_by_class_name('a-list-item')
            for item in list_item:
                temp=str(item.text).strip()
                if temp!='':
                    res.append(temp)
        except:
            return []
        else:
            return res

    def get_pic_url(self):
        try:
            res = []
            pic=self.driver.find_element_by_id('altImages').find_elements_by_tag_name('img')
            for i in pic:
                temp=i.get_attribute('src')
                if 'jpg' in temp and 'icon' not in temp:
                    res.append(prase_pic_size(temp))
        except:
            return []
        else:
            return res

    ##2.21 10 p.m.The video is encrypted by the blob .And solve this problem later.
    def get_video_url(self):
        try:
            src=self.driver.find_element_by_tag_name('video').get_attribute('src')
        except:
            return ''
        else:
            return src
    def get_product_description(self):
        try:
            dption=self.driver.find_element_by_id('productDescription').text
        except:
            try:
                dption_option=self.driver.find_element_by_id('aplus').text
            except:
                return ''
            else:
                return str(dption_option).strip()
        return str(dption).strip()
    def get_prodDetails(self):
        try:
            prodDetails=self.driver.find_element_by_id('prodDetails').text
        except:
            return []
        else:
            res=[]
            textlst=str(prodDetails).split('\n')
            for i in textlst:
                res.append(i.strip())
            return res

    #2.21  12p.m. It is easy to be wrong and the codes need to be refactored.
    #2.22  7p.m. Solved this problem
    def get_remarks(self):
        try:
            remark = self.driver.find_element_by_class_name('card-padding').find_elements_by_tag_name('div')
            time.sleep(5)
        except:
            return []
        else:
            res = []
            for i in remark:
                try:
                    temp = i.get_attribute('data-hook')
                    if temp == 'review-collapsed':
                        res.append(str(i.text).strip())
                except:
                    pass
            return res

    def return_one_shop_information(self,url):
        res=[]
        try:
            self.init_one_shop(url)
            time.sleep(random.randint(5,15))
            res.append(self.get_title())
            res.append(self.get_score())
            res.append(self.get_price_and_delivery())
            res.append(toStr(self.get_list_item()))
            res.append(str(self.get_pic_url()))
              # print(self.get_video_url())
            res.append(str(self.get_product_description()))
            res.append(toStr(self.get_prodDetails()))
            res.append(toStr(self.get_remarks()))
        except:
            return ''
        else:
            res_str=''
            for i in res:
                res_str+=i
                res_str+='&#&' # the data delimiter
            #print(res_str)
            return res_str

    def return_shop_information(self, urls):
        if len(urls)==0:
            return []
        else:
            res=[]
            error_url=[]
            for url in urls:
                try:
                    temp=self.return_one_shop_information(url)
                    print(temp)
                except:
                    error_url.append(url)
                    pass
                else:
                    if temp!=None:
                        res.append(temp)
                time.sleep(random.randint(5,10))
            return res,error_url

    def __del__(self):
        self.driver.close()
# to generate the dataframe
class GetDataFrame():
    def __init__(self,allshop_url,nums=0):
       self.data=[]
       self.allshop_url=allshop_url
       self.spi=SHOPSPIDER()
       self.nums=nums

    def toarr(self,raw_str):
        if raw_str=='': return []
        return raw_str.split('&#&')

    def process(self,url):
        res=self.spi.return_one_shop_information(url)
        if len(res)!=0:
            self.data.append(self.toarr(res))
            #print(res)

    def processForMul(self, threads):
        for i in range(threads, self.nums, self.threads):
            self.process(self.allshop_url[i])

    def multiply_process(self):
        self.threads = 8
        pool = Pool(self.threads)
        pool.map(self.processForMul,range(self.threads))
        pool.close()
        pool.join()
    def return_DataFrame(self):
        summaryDataFrame = pd.DataFrame(self.data)
        return summaryDataFrame
# to format output
class xlsx_saver():
    def __init__(self, df_in, filename='a.xlsx', sheet_name='Sheet1'):
        self.filename = filename
        self.user_def = []
        if os.path.exists(filename):
            self.wb = load_workbook(filename)
            self.sheet = self.wb.create_sheet(sheet_name)
        else:
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
        self.df = df_in.copy()
        self.sheet.append(list(self.df.columns))
        for row in range(0, len(list(self.df.index))):
            for col in range(0, len(list(self.df.columns))):
                self.sheet.cell(row + 2, col + 1).value = self.df.iloc[row, col]

    def set_sheet_name(self, sheet_name):
        self.sheet.title = sheet_name

    def set_filename(self, filename):
        self.filename = filename

    def __get_maxlength(self, series_in):
        series = series_in.fillna('-')  # 填充空值，防止出现nan
        str_list = list(series)
        len_list = []
        for elem in str_list:
            elem_split = list(elem)
            length = 0
            for c in elem_split:
                if ord(c) <= 256:
                    length += 1
                else:
                    length += 2
            len_list.append(length)
        return max(len_list)

    def __auto_width(self):
        cols_list = list(self.df.columns)
        for i in range(0, len(cols_list)):
            col = cols_list[i]
            if col in self.user_def:
                continue
            self.sheet.cell(1, i + 1).font = Font(bold=True)
            letter = chr(i + 65)
            max_len = self.__get_maxlength(self.df[col].astype(str))
            if max_len <= 10:
                self.sheet.column_dimensions[letter].width = 10
            elif max_len <= 50:
                self.sheet.column_dimensions[letter].width = max_len + 1
            else:
                self.sheet.column_dimensions[letter].width = 50
                for cell in self.sheet[letter]:
                    cell.alignment = Alignment(wrap_text=True)

    def set_width(self, col_name, width):
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        self.sheet.column_dimensions[letter].width = width
        self.user_def.append(col_name)

    def set_color(self, col_name, color, rule):
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        for cell in self.sheet[letter]:
            if rule(cell.value):
                cell.fill = PatternFill(fill_type="solid", start_color=color, end_color=color)

    def save(self):
        self.__auto_width()
        self.wb.save(self.filename)

    def set_merge(self, col_name):
        self.user_def.append(col_name)
        index = list(self.df.columns).index(col_name)
        letter = chr(index + 65)
        i = 1
        while True:
            if i >= self.sheet.max_row:
                break
            cell = self.sheet[letter + str(i)]
            j = i + 1
            while True:
                cell_next = self.sheet[letter + str(j)]
                if cell_next.value != cell.value:
                    j -= 1
                    break
                else:
                    j += 1
                if j > self.sheet.max_row:
                    j -= 1
                    break
            if j - i >= 1 and cell.value != '' and cell.value != None:
                msg = '%s%d:%s%d' % (letter, i, letter, j)
                self.sheet.merge_cells(msg)
            self.sheet[letter + str(i)].alignment = Alignment(horizontal='center',
                                                              vertical='top',
                                                              wrap_text=True)
            i = j + 1

def parse_raw_url(raw_url,last_page):
    #Generate all product display pages
    pattern1=re.compile('page=([0-9]+)&')
    pattern2=re.compile('_([0-9]+)')
    res=[]
    for i in range(1,last_page+1):
        raw_temp=re.sub(pattern1,'page='+ str(i)+'&' , raw_url)
        temp=re.sub(pattern2,'_'+ str(i) , raw_temp)
        res.append(temp)
    return res

# To Remove all images size limitation
def prase_pic_size(raw_url):
    p1=re.compile(r'\._SS40_')
    url=re.sub(p1,'',raw_url)
    return url

# array to str
def toStr(numarray):
    res=''
    for i in numarray:
        res+=i
    return res

def writer(num):
    f=open('url.txt','w',encoding='utf-8')
    for i in num :
        f.write(i)
        f.write('\n')
    f.close()

# main fun
def mymain():
    keywords = str(input("Please input a key word "))
    nums = int(input('Please enter the number of products'))
    spi = SPIDER()
    all_page_url = spi.get_allshop_url(keywords)
    lenshop=len(all_page_url)
    if lenshop==0:
        print('Product link not obtained!')
    else:
        print('Already got'+str(lenshop) +'product links!')
        writer(all_page_url)
        if nums>lenshop:
            nums=lenshop
        DataFrame=GetDataFrame(allshop_url=all_page_url,nums=nums)
        DataFrame.multiply_process()
        dataframe = DataFrame.return_DataFrame()
        saver = xlsx_saver(dataframe, keywords + '.xlsx', 'data')
        saver.save()
        print('Sucess!')





if __name__=="__main__":
    mymain()



